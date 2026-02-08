"""
Supabase Ingestion with Dataset Selection
==========================================

Ingest documents from different dataset sources into Supabase.

Usage:
    python ingest_supabase_datasets.py --dataset OLD
    python ingest_supabase_datasets.py --dataset COMBINED
    python ingest_supabase_datasets.py --dataset NEW
"""

import os
import json
import time
import argparse
from pathlib import Path
import sys

# Add src to path
sys.path.append('src')

from config_datasets import get_dataset_config, list_datasets, DatasetType
from embed import embed_texts
import os
from dotenv import load_dotenv
from PyPDF2 import PdfReader
from bs4 import BeautifulSoup
import re
from docx import Document
import openpyxl

load_dotenv()

# Supabase credentials - Use SERVICE_KEY for ingestion (admin access)
SUPABASE_URL = os.getenv("SUPABASE_URL")
SUPABASE_KEY = os.getenv("SUPABASE_SERVICE_KEY") or os.getenv("SUPABASE_KEY")

if not SUPABASE_URL:
    print("[FAIL] ERROR: SUPABASE_URL not found in .env")
    sys.exit(1)
if not SUPABASE_KEY:
    print("[FAIL] ERROR: SUPABASE_SERVICE_KEY or SUPABASE_KEY not found in .env")
    sys.exit(1)


def load_documents_from_dir(directory):
    """Load documents from a directory"""
    documents = []
    directory = Path(directory)
    
    if not directory.exists():
        print(f"[WARN]  Directory not found: {directory}")
        return documents
    
    # Handle old dataset (PDF, DOCX, XLSX files structure)
    if 'scraped_dpmptsp' in str(directory):
        files_dir = directory / 'files'
        if files_dir.exists():
            # Load PDF files
            print(f"  [FILE] Loading PDF files from {files_dir}...")
            for pdf_file in files_dir.glob('*.pdf'):
                try:
                    reader = PdfReader(pdf_file)
                    text_content = []
                    for page in reader.pages:
                        text_content.append(page.extract_text())
                    text = ' '.join(text_content)
                    text = ' '.join(text.split())  # Normalize whitespace
                    
                    if text.strip():
                        documents.append({
                            'title': pdf_file.stem,
                            'content': text[:5000],  # Limit to 5000 chars
                            'source': 'scraped_dpmptsp',
                            'file': pdf_file.name,
                        })
                except Exception as e:
                    pass  # Silent fail for corrupted PDFs
            
            # Load DOCX files
            print(f"  [FILE] Loading DOCX files from {files_dir}...")
            for docx_file in files_dir.glob('*.docx'):
                try:
                    doc = Document(docx_file)
                    text = '\n'.join([para.text for para in doc.paragraphs])
                    text = ' '.join(text.split())  # Normalize whitespace
                    
                    if text.strip():
                        documents.append({
                            'title': docx_file.stem,
                            'content': text[:5000],  # Limit to 5000 chars
                            'source': 'scraped_dpmptsp',
                            'file': docx_file.name,
                        })
                except Exception as e:
                    print(f"    [FAIL] Error reading {docx_file.name}: {e}")
            
            # Load XLSX files
            print(f"  [FILE] Loading XLSX files from {files_dir}...")
            for xlsx_file in files_dir.glob('*.xlsx'):
                try:
                    workbook = openpyxl.load_workbook(xlsx_file)
                    text_parts = []
                    for sheet in workbook.sheetnames:
                        ws = workbook[sheet]
                        for row in ws.iter_rows(values_only=True):
                            text_parts.extend([str(cell) for cell in row if cell])
                    text = ' '.join(text_parts)
                    text = ' '.join(text.split())  # Normalize whitespace
                    
                    if text.strip():
                        documents.append({
                            'title': xlsx_file.stem,
                            'content': text[:5000],  # Limit to 5000 chars
                            'source': 'scraped_dpmptsp',
                            'file': xlsx_file.name,
                        })
                except Exception as e:
                    print(f"    [FAIL] Error reading {xlsx_file.name}: {e}")
            
            # Load DOC files (legacy format)
            print(f"  [FILE] Loading DOC files from {files_dir}...")
            for doc_file in files_dir.glob('*.doc'):
                try:
                    # DOC files are binary and harder to parse, try basic extraction
                    with open(doc_file, 'rb') as f:
                        content = f.read()
                        # Very basic text extraction from binary
                        text = ''.join(chr(b) for b in content if 32 <= b < 127)
                    text = ' '.join(text.split())
                    
                    if text.strip() and len(text) > 50:
                        documents.append({
                            'title': doc_file.stem,
                            'content': text[:5000],
                            'source': 'scraped_dpmptsp',
                            'file': doc_file.name,
                        })
                except Exception as e:
                    pass  # Silent fail for DOC files
            
            # Also check for HTML files
            print(f"  [FILE] Loading HTML files from {files_dir}...")
            for html_file in files_dir.glob('*.html'):
                try:
                    with open(html_file, 'r', encoding='utf-8') as f:
                        content = f.read()
                        # Simple HTML cleaning (remove tags)
                        text = re.sub('<[^<]+?>', '', content)
                        text = ' '.join(text.split())  # Normalize whitespace
                        
                        if text.strip():
                            documents.append({
                                'title': html_file.stem,
                                'content': text[:5000],  # Limit to 5000 chars
                                'source': 'scraped_dpmptsp',
                                'file': html_file.name,
                            })
                except Exception as e:
                    print(f"    [FAIL] Error reading {html_file.name}: {e}")


    
    # Handle new dataset (NDJSON files)
    else:
        for json_file in directory.rglob('*.ndjson'):
            print(f"  [FILE] Loading {json_file}...")
            try:
                with open(json_file, 'r', encoding='utf-8') as f:
                    for line in f:
                        if line.strip():
                            doc = json.loads(line)
                            if isinstance(doc, dict):
                                documents.append({
                                    'title': doc.get('title', 'Untitled'),
                                    'content': doc.get('text', doc.get('content', ''))[:5000],
                                    'source': 'data_oss',
                                    'category': doc.get('category', 'General'),
                                })
            except Exception as e:
                print(f"    [FAIL] Error reading {json_file}: {e}")
        
        # Also check for regular JSON files
        for json_file in directory.rglob('*.json'):
            if json_file.name != 'crawl_summary.json':
                print(f"  [FILE] Loading {json_file}...")
                try:
                    with open(json_file, 'r', encoding='utf-8') as f:
                        data = json.load(f)
                        if isinstance(data, list):
                            documents.extend([
                                {
                                    'title': item.get('title', 'Untitled'),
                                    'content': item.get('text', item.get('content', ''))[:5000],
                                    'source': 'data_oss',
                                }
                                for item in data if isinstance(item, dict)
                            ])
                except Exception as e:
                    print(f"    [FAIL] Error reading {json_file}: {e}")
    
    return documents


def ingest_dataset(dataset_type: DatasetType):
    """Ingest a specific dataset into Supabase"""
    config = get_dataset_config(dataset_type)
    
    print("\n" + "="*70)
    print(f"INGESTING {dataset_type} DATASET")
    print("="*70)
    print(f"Table: {config.table_name}")
    print(f"Description: {config.description}")
    print(f"Source directories: {config.source_dirs}")
    print()
    
    # Load all documents
    print("\n Loading documents from sources...")
    all_documents = []
    for source_dir in config.source_dirs:
        print(f"\n  Processing {source_dir}...")
        docs = load_documents_from_dir(source_dir)
        all_documents.extend(docs)
        print(f"    [OK] Loaded {len(docs)} documents")
    
    total_docs = len(all_documents)
    print(f"\n[OK] Total documents loaded: {total_docs}")
    
    if total_docs == 0:
        print("[FAIL] No documents found to ingest!")
        return False
    
    # Initialize Supabase
    print("\n[CONNECT] Initializing Supabase connection...")
    from supabase import create_client
    supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
    
    # Check if table exists
    print(f"[STATS] Checking if table '{config.table_name}' exists...")
    try:
        result = supabase.table(config.table_name).select('id').limit(1).execute()
        print(f"   [OK] Table exists")
    except Exception as e:
        print(f"   [WARN]  Table may not exist yet: {e}")
        print(f"   Will create on first insert")
    
    # Ingest documents
    print(f"\n[START] Ingesting {total_docs} documents...")
    batch_size = 10
    successful = 0
    failed = 0
    
    for i in range(0, total_docs, batch_size):
        batch = all_documents[i:i+batch_size]
        batch_num = (i // batch_size) + 1
        total_batches = (total_docs + batch_size - 1) // batch_size
        
        print(f"\n  Batch {batch_num}/{total_batches} ({len(batch)} documents)...")
        
        # Generate embeddings for batch
        batch_texts = [doc['content'] for doc in batch]
        try:
            batch_embeddings = embed_texts(batch_texts)
        except Exception as e:
            print(f"    [FAIL] Error generating embeddings for batch: {e}")
            failed += len(batch)
            continue
        
        for j, (doc, embedding) in enumerate(zip(batch, batch_embeddings), 1):
            try:
                # Prepare document for Supabase
                doc_data = {
                    'title': doc['title'],
                    'content': doc['content'],
                    'source': doc.get('source', 'unknown'),
                    'embedding': embedding,
                }
                
                # Add optional fields
                if 'category' in doc:
                    doc_data['category'] = doc['category']
                if 'file' in doc:
                    doc_data['file'] = doc['file']
                
                # Insert into Supabase
                result = supabase.table(config.table_name).insert(doc_data).execute()
                successful += 1
                print(f"    [OK] {j}. {doc['title'][:50]}")
                
            except Exception as e:
                failed += 1
                print(f"    [FAIL] {j}. {doc['title'][:50]}: {str(e)[:50]}")
        
        # Rate limiting
        if batch_num < total_batches:
            time.sleep(1)
    
    # Summary
    print("\n" + "="*70)
    print("INGESTION COMPLETE")
    print("="*70)
    print(f"[OK] Successful: {successful}")
    print(f"[FAIL] Failed: {failed}")
    print(f"[STATS] Total: {total_docs}")
    print(f" Table: {config.table_name}")
    print("="*70 + "\n")
    
    return failed == 0


def main():
    parser = argparse.ArgumentParser(
        description='Ingest documents into Supabase from different datasets'
    )
    parser.add_argument(
        '--dataset',
        type=str,
        choices=['NEW', 'OLD', 'COMBINED'],
        default='NEW',
        help='Dataset to ingest (NEW, OLD, or COMBINED)'
    )
    parser.add_argument(
        '--list',
        action='store_true',
        help='List available datasets'
    )
    
    args = parser.parse_args()
    
    if args.list:
        list_datasets()
        return
    
    # Ingest the selected dataset
    success = ingest_dataset(args.dataset)
    exit(0 if success else 1)


if __name__ == "__main__":
    main()
