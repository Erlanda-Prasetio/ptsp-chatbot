"""
Batch ingest script for DPMPTSP scraped data
Processes all scraped text files and downloaded documents
"""
import os
import sys
from pathlib import Path
import time
from typing import List, Dict, Any
import json
from concurrent.futures import ThreadPoolExecutor, as_completed
import logging

# Add src to path
sys.path.append('src')

from enhanced_rag import EnhancedRAG
from embed import embed_texts
from vector_store_supabase_rest import SupabaseRestVectorStore
import PyPDF2
import openpyxl
import xlrd
import csv

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class DPMPTSPDataIngestor:
    def __init__(self):
        self.store = SupabaseRestVectorStore()
        self.scraped_dir = Path("data/scraped_dpmptsp")
        self.pages_dir = self.scraped_dir / "pages"
        self.files_dir = self.scraped_dir / "files"
        
        # Chunk settings - larger chunks for better context and accuracy
        self.chunk_size = 800  # Increased from 500 for better context
        self.chunk_overlap = 100  # Increased overlap for continuity
        
        # Statistics
        self.total_files = 0
        self.processed_files = 0
        self.total_chunks = 0
        self.failed_files = []

    def clean_text(self, text: str) -> str:
        """Clean and normalize text"""
        if not text:
            return ""
        
        # Remove excessive whitespace
        text = ' '.join(text.split())
        
        # Remove control characters
        text = ''.join(char for char in text if ord(char) >= 32 or char in '\n\t')
        
        return text.strip()

    def create_chunks(self, text: str, filename: str, metadata: Dict[str, Any]) -> List[Dict[str, Any]]:
        """Split text into chunks with metadata"""
        if not text or len(text.strip()) < 50:
            return []
        
        words = text.split()
        chunks = []
        
        for i in range(0, len(words), self.chunk_size - self.chunk_overlap):
            chunk_words = words[i:i + self.chunk_size]
            chunk_text = ' '.join(chunk_words)
            
            chunk_metadata = {
                **metadata,
                'filename': filename,
                'chunk_index': len(chunks),
                'word_count': len(chunk_words),
                'source': f"DPMPTSP_{filename}"
            }
            
            chunks.append({
                'content': self.clean_text(chunk_text),
                'metadata': chunk_metadata
            })
        
        return chunks

    def process_text_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process a text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            
            # Extract metadata from the file if it exists
            metadata = {
                'file_type': 'webpage_text',
                'source_url': 'unknown',
                'title': 'unknown'
            }
            
            # Try to extract metadata from file header
            lines = content.split('\n')
            for line in lines[:10]:  # Check first 10 lines
                if line.startswith('URL:'):
                    metadata['source_url'] = line.replace('URL:', '').strip()
                elif line.startswith('Title:'):
                    metadata['title'] = line.replace('Title:', '').strip()
                elif line.startswith('=' * 20):  # Content separator
                    break
            
            # Remove metadata from content
            if '=' * 20 in content:
                content = content.split('=' * 20, 1)[1].strip()
            
            return self.create_chunks(content, file_path.name, metadata)
            
        except Exception as e:
            logger.error(f"❌ Error processing text file {file_path}: {e}")
            return []

    def process_pdf_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process a PDF file"""
        try:
            text = ""
            with open(file_path, 'rb') as f:
                pdf_reader = PyPDF2.PdfReader(f)
                for page_num, page in enumerate(pdf_reader.pages):
                    try:
                        page_text = page.extract_text()
                        if page_text:
                            text += f"\n--- Page {page_num + 1} ---\n{page_text}"
                    except Exception as e:
                        logger.warning(f"⚠️  Error extracting page {page_num + 1} from {file_path}: {e}")
            
            metadata = {
                'file_type': 'pdf',
                'source_url': 'dpmptsp_download',
                'title': file_path.stem,
                'total_pages': len(pdf_reader.pages) if 'pdf_reader' in locals() else 0
            }
            
            return self.create_chunks(text, file_path.name, metadata)
            
        except Exception as e:
            logger.error(f"❌ Error processing PDF {file_path}: {e}")
            return []

    def process_excel_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process Excel files"""
        try:
            text = ""
            metadata = {
                'file_type': 'excel',
                'source_url': 'dpmptsp_download',
                'title': file_path.stem
            }
            
            if file_path.suffix.lower() in ['.xlsx', '.xlsm']:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    text += f"\n--- Sheet: {sheet_name} ---\n"
                    
                    for row in sheet.iter_rows(values_only=True):
                        row_text = ' | '.join(str(cell) if cell is not None else '' for cell in row)
                        if row_text.strip():
                            text += row_text + '\n'
            
            elif file_path.suffix.lower() in ['.xls']:
                workbook = xlrd.open_workbook(file_path)
                for sheet_idx in range(workbook.nsheets):
                    sheet = workbook.sheet_by_index(sheet_idx)
                    text += f"\n--- Sheet: {sheet.name} ---\n"
                    
                    for row_idx in range(sheet.nrows):
                        row_values = sheet.row_values(row_idx)
                        row_text = ' | '.join(str(cell) for cell in row_values)
                        if row_text.strip():
                            text += row_text + '\n'
            
            return self.create_chunks(text, file_path.name, metadata)
            
        except Exception as e:
            logger.error(f"❌ Error processing Excel file {file_path}: {e}")
            return []

    def process_csv_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process CSV files"""
        try:
            text = ""
            with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
                csv_reader = csv.reader(f)
                for row in csv_reader:
                    row_text = ' | '.join(row)
                    if row_text.strip():
                        text += row_text + '\n'
            
            metadata = {
                'file_type': 'csv',
                'source_url': 'dpmptsp_download',
                'title': file_path.stem
            }
            
            return self.create_chunks(text, file_path.name, metadata)
            
        except Exception as e:
            logger.error(f"❌ Error processing CSV file {file_path}: {e}")
            return []

    def process_file(self, file_path: Path) -> List[Dict[str, Any]]:
        """Process a single file based on its extension"""
        logger.info(f"📄 Processing: {file_path.name}")
        
        file_ext = file_path.suffix.lower()
        
        if file_ext == '.txt':
            return self.process_text_file(file_path)
        elif file_ext == '.pdf':
            return self.process_pdf_file(file_path)
        elif file_ext in ['.xlsx', '.xls', '.xlsm']:
            return self.process_excel_file(file_path)
        elif file_ext == '.csv':
            return self.process_csv_file(file_path)
        else:
            logger.warning(f"⚠️  Unsupported file type: {file_ext}")
            return []

    def embed_and_store_chunks(self, chunks: List[Dict[str, Any]], batch_size: int = 10) -> int:
        """Embed chunks and store in vector database"""
        if not chunks:
            return 0
        
        stored_count = 0
        
        # Process in batches
        for i in range(0, len(chunks), batch_size):
            batch = chunks[i:i + batch_size]
            batch_texts = [chunk['content'] for chunk in batch]
            
            try:
                # Generate embeddings
                logger.info(f"🔮 Embedding batch {i//batch_size + 1} ({len(batch)} chunks)")
                embeddings = embed_texts(batch_texts)
                
                # Store in database
                chunks_with_embeddings = []
                for j, (chunk, embedding) in enumerate(zip(batch, embeddings)):
                    chunks_with_embeddings.append({
                        'content': chunk['content'],
                        'metadata': chunk['metadata'],
                        'embedding': embedding
                    })
                
                try:
                    success = self.store.add_chunks(chunks_with_embeddings)
                    if success:
                        stored_count += len(chunks_with_embeddings)
                        logger.info(f"✅ Stored batch {i//batch_size + 1}: {len(batch)} chunks")
                    else:
                        logger.error(f"❌ Failed to store batch {i//batch_size + 1}")
                except Exception as e:
                    logger.error(f"❌ Error storing batch {i//batch_size + 1}: {e}")
                
            except Exception as e:
                logger.error(f"❌ Error processing batch {i//batch_size + 1}: {e}")
        
        return stored_count

    def ingest_all_data(self, max_workers: int = 2):
        """Ingest all scraped data"""
        logger.info("🚀 Starting DPMPTSP data ingestion...")
        
        # Get all files to process
        files_to_process = []
        
        # Add text files from pages
        if self.pages_dir.exists():
            text_files = list(self.pages_dir.glob("*.txt"))
            files_to_process.extend(text_files)
            logger.info(f"📄 Found {len(text_files)} text files")
        
        # Add downloaded files
        if self.files_dir.exists():
            for pattern in ['*.pdf', '*.xlsx', '*.xls', '*.csv', '*.txt']:
                pattern_files = list(self.files_dir.glob(pattern))
                files_to_process.extend(pattern_files)
            logger.info(f"📁 Found {len(files_to_process) - len(text_files) if self.pages_dir.exists() else len(files_to_process)} downloaded files")
        
        self.total_files = len(files_to_process)
        logger.info(f"📊 Total files to process: {self.total_files}")
        
        if not files_to_process:
            logger.warning("⚠️  No files found to process!")
            return
        
        # Process files
        all_chunks = []
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(self.process_file, file_path): file_path for file_path in files_to_process}
            
            for future in as_completed(futures):
                file_path = futures[future]
                try:
                    chunks = future.result()
                    if chunks:
                        all_chunks.extend(chunks)
                        logger.info(f"✅ Processed {file_path.name}: {len(chunks)} chunks")
                    else:
                        logger.warning(f"⚠️  No chunks extracted from {file_path.name}")
                    
                    self.processed_files += 1
                    progress = (self.processed_files / self.total_files) * 100
                    logger.info(f"📊 Progress: {self.processed_files}/{self.total_files} files ({progress:.1f}%)")
                    
                except Exception as e:
                    logger.error(f"❌ Failed to process {file_path}: {e}")
                    self.failed_files.append(str(file_path))
        
        # Store all chunks
        if all_chunks:
            logger.info(f"🔮 Storing {len(all_chunks)} chunks in vector database...")
            stored_count = self.embed_and_store_chunks(all_chunks)
            self.total_chunks = stored_count
            
            # Get total count in database
            total_in_db = self.store.get_count()
            
            logger.info("🎉 DPMPTSP Data Ingestion Complete!")
            logger.info(f"📄 Files processed: {self.processed_files}/{self.total_files}")
            logger.info(f"📦 New chunks added: {stored_count}")
            logger.info(f"📊 Total chunks in database: {total_in_db}")
            logger.info(f"❌ Failed files: {len(self.failed_files)}")
            
            if self.failed_files:
                logger.info("Failed files:")
                for file in self.failed_files:
                    logger.info(f"  - {file}")
        else:
            logger.warning("⚠️  No chunks were created from the processed files!")

def main():
    """Main function"""
    ingestor = DPMPTSPDataIngestor()
    
    try:
        ingestor.ingest_all_data(max_workers=2)
    except KeyboardInterrupt:
        print("\n🛑 Ingestion interrupted by user")
    except Exception as e:
        print(f"❌ Error during ingestion: {e}")

if __name__ == "__main__":
    main()
